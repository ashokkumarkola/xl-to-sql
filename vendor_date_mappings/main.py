"""Convert selected columns from an Excel worksheet into a delimited text file."""

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import TextIO

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


# ==================================================
# Exceptions
# ==================================================

class ExcelToTxtError(Exception):
    """Base exception for Excel-to-TXT conversion errors."""


class InputFileError(ExcelToTxtError):
    """Raised when the input Excel file is invalid or unavailable."""


class SheetNotFoundError(ExcelToTxtError):
    """Raised when the requested worksheet does not exist."""


class ColumnNotFoundError(ExcelToTxtError):
    """Raised when a required column does not exist."""


class DataConversionError(ExcelToTxtError):
    """Raised when Excel data cannot be converted."""


# ==================================================
# Configuration
# ==================================================

INPUT_FILE = Path("data/Lido Mapping June 18.xlsx")
OUTPUT_FILE = Path("Lido Mapping June 18.txt")

SHEET_NAME = "in"

INCLUDE_HEADER = False

SELECTED_COLUMNS = [
    # "Invoice - Vendor",
    "NextGen - Vendor",
    "Invoice - Date Format",
]

COLUMN_SEPARATOR = " | "

SKIP_EMPTY_ROWS = True


# ==================================================
# Data
# ==================================================

@dataclass
class ConversionStats:
    """Row-level counters collected while writing the output file."""

    total_rows: int = 0
    written_rows: int = 0
    empty_rows: int = 0


# ==================================================
# Validation helpers
# ==================================================

def validate_input_file(input_file: Path) -> None:
    """Ensure the input path exists, is a file, and has a supported extension."""

    if not input_file.exists():
        raise InputFileError(f"Input Excel file not found: '{input_file}'")

    if not input_file.is_file():
        raise InputFileError(f"Input path is not a file: '{input_file}'")

    if input_file.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise InputFileError(
            f"Unsupported Excel file type: '{input_file.suffix}'. "
            f"Expected .xlsx or .xlsm."
        )


def validate_output_directory(output_file: Path) -> None:
    """Ensure the output file's parent directory exists."""

    output_directory = output_file.parent

    if not output_directory.exists():
        raise ExcelToTxtError(
            f"Output directory does not exist: '{output_directory}'"
        )


# ==================================================
# Workbook / worksheet helpers
# ==================================================

def open_workbook(input_file: Path) -> Workbook:
    """Load an Excel workbook, translating low-level errors into InputFileError."""

    try:
        return load_workbook(input_file, data_only=True)
    except InvalidFileException as exc:
        raise InputFileError(
            f"Unable to read Excel file '{input_file}'. "
            f"The file may be corrupted or invalid."
        ) from exc
    except PermissionError as exc:
        raise InputFileError(
            f"Permission denied while reading Excel file '{input_file}'."
        ) from exc
    except OSError as exc:
        raise InputFileError(
            f"Unable to open Excel file '{input_file}': {exc}"
        ) from exc


def get_worksheet(workbook: Workbook, sheet_name: str, input_file: Path) -> Worksheet:
    """Return the requested worksheet, raising SheetNotFoundError if missing."""

    if sheet_name not in workbook.sheetnames:
        available_sheets = ", ".join(f"'{name}'" for name in workbook.sheetnames)

        raise SheetNotFoundError(
            f"Worksheet '{sheet_name}' was not found in '{input_file.name}'.\n"
            f"Available worksheets: {available_sheets}\n"
            f"Please update SHEET_NAME in the configuration."
        )

    sheet = workbook[sheet_name]

    if sheet.max_row < 1:
        raise ExcelToTxtError(
            f"Worksheet '{sheet_name}' is empty. "
            f"Expected the first row to contain column headers."
        )

    return sheet


def read_headers(sheet: Worksheet, sheet_name: str) -> list[str]:
    """Read and normalize the header row (row 1) of a worksheet."""

    headers = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in sheet[1]
    ]

    if not any(headers):
        raise ExcelToTxtError(
            f"Worksheet '{sheet_name}' does not contain "
            f"any column headers in the first row."
        )

    return headers


def resolve_column_indexes(
    headers: list[str],
    selected_columns: list[str],
    sheet_name: str,
) -> dict[str, int]:
    """Map each selected column name to its position in the header row."""

    column_indexes: dict[str, int] = {}

    for column_name in selected_columns:

        if column_name not in headers:
            available_columns = ", ".join(
                f"'{column}'" for column in headers if column
            )

            raise ColumnNotFoundError(
                f"Required column '{column_name}' was not found "
                f"in worksheet '{sheet_name}'.\n"
                f"Available columns: {available_columns}\n"
                f"Please check SELECTED_COLUMNS."
            )

        column_indexes[column_name] = headers.index(column_name)

    return column_indexes


# ==================================================
# Row processing helpers
# ==================================================

def extract_row_values(
    row: tuple,
    selected_columns: list[str],
    column_indexes: dict[str, int],
    row_number: int,
) -> list[str]:
    """Pull and normalize the selected column values out of one Excel row."""

    values = []

    for column_name in selected_columns:
        column_index = column_indexes[column_name]

        try:
            value = row[column_index]
        except IndexError as exc:
            raise DataConversionError(
                f"Unable to read column '{column_name}' at Excel row {row_number}."
            ) from exc

        values.append("" if value is None else str(value).strip())

    return values


def is_empty_row(values: list[str]) -> bool:
    """A row counts as empty when every selected value is blank."""

    return not any(values)


def write_rows(
    txt_file: TextIO,
    sheet: Worksheet,
    selected_columns: list[str],
    column_indexes: dict[str, int],
    column_separator: str,
    skip_empty_rows: bool,
) -> ConversionStats:
    """Write data rows to the output file, logging the outcome of each row."""

    stats = ConversionStats()

    for row_number, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        stats.total_rows += 1

        values = extract_row_values(row, selected_columns, column_indexes, row_number)

        if is_empty_row(values):
            stats.empty_rows += 1

            if skip_empty_rows:
                logger.info("  [SKIP]  Row %d: empty, skipped.", row_number)
                continue

            logger.info("  [EMPTY] Row %d: empty, kept.", row_number)

        stats.written_rows += 1
        logger.info(
            "  [OK]    Row %d: %s", row_number, column_separator.join(values)
        )

        txt_file.write(column_separator.join(values) + "\n")

    return stats


def log_summary(
    input_file: Path,
    output_file: Path,
    sheet_name: str,
    stats: ConversionStats,
    skip_empty_rows: bool,
) -> None:
    """Log a final summary of the conversion run."""

    logger.info(
        "\nConversion successful.\n"
        "Input       : %s\n"
        "Sheet       : %s\n"
        "Output      : %s\n"
        "Rows scanned: %d\n"
        "Rows written: %d\n"
        "Empty rows  : %d%s",
        input_file,
        sheet_name,
        output_file,
        stats.total_rows,
        stats.written_rows,
        stats.empty_rows,
        " (skipped)" if skip_empty_rows else " (kept)",
    )


# ==================================================
# Orchestration
# ==================================================

def excel_to_txt(
    input_file: Path,
    output_file: Path,
    sheet_name: str,
    selected_columns: list[str],
    include_header: bool = False,
    column_separator: str = " | ",
    skip_empty_rows: bool = True,
) -> ConversionStats:
    """Convert selected columns of an Excel worksheet into a delimited text file."""

    validate_input_file(input_file)
    validate_output_directory(output_file)

    workbook = open_workbook(input_file)

    try:
        sheet = get_worksheet(workbook, sheet_name, input_file)
        headers = read_headers(sheet, sheet_name)
        column_indexes = resolve_column_indexes(headers, selected_columns, sheet_name)

        logger.info("Reading rows from sheet '%s'...", sheet_name)

        try:
            with output_file.open("w", encoding="utf-8") as txt_file:

                if include_header:
                    txt_file.write(column_separator.join(selected_columns) + "\n")

                stats = write_rows(
                    txt_file,
                    sheet,
                    selected_columns,
                    column_indexes,
                    column_separator,
                    skip_empty_rows,
                )

        except PermissionError as exc:
            raise ExcelToTxtError(
                f"Permission denied while writing output file '{output_file}'."
            ) from exc
        except OSError as exc:
            raise ExcelToTxtError(
                f"Unable to write output file '{output_file}': {exc}"
            ) from exc

    finally:
        workbook.close()

    log_summary(input_file, output_file, sheet_name, stats, skip_empty_rows)

    return stats


# ==================================================
# Main
# ==================================================

def configure_logging(level: int = logging.INFO) -> None:
    """Configure a simple console logger for CLI usage."""

    logging.basicConfig(level=level, format="%(message)s")


def main() -> None:
    configure_logging()

    try:
        excel_to_txt(
            input_file=INPUT_FILE,
            output_file=OUTPUT_FILE,
            sheet_name=SHEET_NAME,
            selected_columns=SELECTED_COLUMNS,
            include_header=INCLUDE_HEADER,
            column_separator=COLUMN_SEPARATOR,
            skip_empty_rows=SKIP_EMPTY_ROWS,
        )

    except ExcelToTxtError as exc:
        logger.error("\nERROR: %s\n", exc)

    except KeyboardInterrupt:
        logger.error("\nERROR: Operation cancelled by user.")

    except Exception:
        logger.exception("\nUNEXPECTED ERROR:")
        raise


if __name__ == "__main__":
    main()
